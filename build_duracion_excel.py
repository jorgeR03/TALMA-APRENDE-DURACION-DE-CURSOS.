import json, re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

with open('cursos_2026_tiempos.json', encoding='utf-8') as f:
    cursos = json.load(f)

def parse_minutos(tiempos):
    if not tiempos:
        return None
    t = tiempos[0]
    m = re.match(r'(\d+)\s*h,\s*(\d+)\s*min,\s*(\d+)\s*seg', t)
    if not m:
        return None
    h, mi, s = (int(x) for x in m.groups())
    total = h * 60 + mi + round(s / 60)
    return total

rows = []
for curso, e in cursos.items():
    minutos = parse_minutos(e['tiempos'])
    inst = ', '.join(e['instituciones'])
    rows.append(dict(curso=curso, minutos=minutos, institucion=inst, n=e['n']))

# Pendientes por completar primero, luego alfabetico dentro de cada grupo
rows.sort(key=lambda r: (r['minutos'] is not None, r['curso'].upper()))

OUT = 'Tiempo de Duracion de Cursos.xlsx'
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Duracion Cursos'
ws.sheet_view.showGridLines = False

NAVY = 'FF12233D'
NAVY_2 = 'FF1B3A63'
YELLOW = 'FFFFF2CC'
YELLOW_TXT = 'FF8A6D00'
GOOD = 'FFE7F5E8'

ws.merge_cells('A1:E2')
c = ws['A1']
c.value = '  TIEMPO DE DURACIÓN DE CURSOS — 2026'
c.font = Font(size=18, bold=True, color='FFFFFFFF')
c.alignment = Alignment(vertical='center')
for row_ in ws['A1:E2']:
    for cell in row_:
        cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 22

ws.merge_cells('A3:E3')
c = ws['A3']
n_total = len(rows)
n_con = sum(1 for r in rows if r['minutos'] is not None)
n_sin = n_total - n_con
c.value = f'  Talma Servicios Aeroportuarios · {n_total} cursos con "2026" en el reporteador global · {n_con} con duración registrada · {n_sin} pendientes por completar manualmente'
c.font = Font(size=10, italic=True, color='FFFFFFFF')
c.alignment = Alignment(vertical='center')
for row_ in ws['A3:E3']:
    for cell in row_:
        cell.fill = PatternFill(start_color=NAVY_2, end_color=NAVY_2, fill_type='solid')
ws.row_dimensions[3].height = 18
ws.row_dimensions[4].height = 8

headers = ['CURSO', 'CLIENTE / INSTITUCIÓN', 'DURACIÓN (min)', 'DURACIÓN', 'REGISTROS EN 2026']
header_row = 5
for j, h in enumerate(headers):
    cell = ws.cell(row=header_row, column=1 + j, value=h)
    cell.font = Font(bold=True, color='FFFFFFFF', size=10)
    cell.fill = PatternFill(start_color=NAVY_2, end_color=NAVY_2, fill_type='solid')
    cell.alignment = Alignment(horizontal='center' if j > 0 else 'left', vertical='center')

thin = Side(style='thin', color='FFD8DCE3')
box = Border(left=thin, right=thin, top=thin, bottom=thin)

r = header_row + 1
for row_data in rows:
    minutos = row_data['minutos']
    texto = ''
    if minutos is not None:
        hh, mm = divmod(minutos, 60)
        if hh and mm:
            texto = f'{hh} h {mm} min'
        elif hh:
            texto = f'{hh} h'
        else:
            texto = f'{mm} min'
    ws.cell(row=r, column=1, value=row_data['curso'])
    ws.cell(row=r, column=2, value=row_data['institucion'])
    ws.cell(row=r, column=3, value=minutos)
    ws.cell(row=r, column=4, value=texto)
    ws.cell(row=r, column=5, value=row_data['n'])
    for col in range(1, 6):
        cell = ws.cell(row=r, column=col)
        cell.border = box
        if col >= 3:
            cell.alignment = Alignment(horizontal='center')
        if minutos is None:
            cell.fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')
            if col in (3, 4):
                cell.font = Font(color=YELLOW_TXT, bold=True)
                if col == 4:
                    ws.cell(row=r, column=4, value='POR COMPLETAR')
    r += 1
last_row = r - 1

ws.column_dimensions['A'].width = 62
ws.column_dimensions['B'].width = 34
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 16

tab = Table(displayName='DuracionCursos', ref=f'A{header_row}:E{last_row}')
tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
ws.add_table(tab)
ws.freeze_panes = f'A{header_row + 1}'

wb.save(OUT)
print('Guardado:', OUT)
print('Total cursos:', n_total, '| con duracion:', n_con, '| pendientes:', n_sin)
