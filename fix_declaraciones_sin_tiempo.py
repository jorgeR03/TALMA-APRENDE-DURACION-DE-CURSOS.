import openpyxl
from openpyxl.styles import PatternFill, Font

FILE = 'Tiempo de Duracion de Cursos.xlsx'
GRAY = 'FFF0F0EE'
GRAY_TXT = 'FF6B7280'

wb = openpyxl.load_workbook(FILE)
ws = wb['Duracion Cursos']

targets = ['DECLARACION DE CONFLICTO DE INTERES', 'DECLARACION DE PAGOS INDEBIDOS']
changed = []
for r in range(6, ws.max_row + 1):
    curso = ws.cell(row=r, column=1).value
    if not curso:
        continue
    cu = curso.upper()
    if any(t in cu for t in targets):
        ws.cell(row=r, column=3, value='Sin tiempo')
        ws.cell(row=r, column=4, value='Sin tiempo')
        for col in (1, 2, 3, 4, 5):
            cell = ws.cell(row=r, column=col)
            cell.fill = PatternFill(start_color=GRAY, end_color=GRAY, fill_type='solid')
            if col in (3, 4):
                cell.font = Font(color=GRAY_TXT, italic=True)
        changed.append(curso)

wb.save(FILE)
print('Filas actualizadas:', len(changed))
for c in changed:
    print(' -', c)
