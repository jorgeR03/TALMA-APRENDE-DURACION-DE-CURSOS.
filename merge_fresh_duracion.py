import json, re
import openpyxl

FILE = 'Tiempo de Duracion de Cursos.xlsx'

with open('cursos_2026_tiempos_fresh.json', encoding='utf-8') as f:
    fresh_raw = json.load(f)
fresh = {curso.strip(): e for curso, e in fresh_raw.items()}

def parse_minutos(tiempos):
    if not tiempos:
        return None
    t = tiempos[0]
    m = re.match(r'(\d+)\s*h,\s*(\d+)\s*min,\s*(\d+)\s*seg', t)
    if not m:
        return None
    h, mi, s = (int(x) for x in m.groups())
    return h * 60 + mi + round(s / 60)

fresh_min = {curso: parse_minutos(e['tiempos']) for curso, e in fresh.items()}
fresh_inst = {curso: ', '.join(e['instituciones']) for curso, e in fresh.items()}
fresh_n = {curso: e['n'] for curso, e in fresh.items()}

wb = openpyxl.load_workbook(FILE, data_only=True)
ws = wb['Duracion Cursos']
current = {}
current_order = []
for r in range(6, ws.max_row + 1):
    curso = ws.cell(row=r, column=1).value
    if not curso:
        continue
    current[curso] = dict(
        inst=ws.cell(row=r, column=2).value,
        dur=ws.cell(row=r, column=3).value,
        n=ws.cell(row=r, column=5).value,
    )
    current_order.append(curso)

current_set = set(current)
fresh_set = set(fresh)

nuevos = sorted(fresh_set - current_set)
retirados = sorted(current_set - fresh_set)
cambios_num = []

for curso in current_order:
    if curso not in fresh_set:
        continue
    cur = current[curso]
    if isinstance(cur['dur'], (int, float)):
        new_val = fresh_min.get(curso)
        if new_val is not None and new_val != cur['dur']:
            cambios_num.append((curso, cur['dur'], new_val))
            current[curso]['dur'] = new_val

for curso in nuevos:
    current[curso] = dict(
        inst=fresh_inst.get(curso, ''),
        dur='Nuevo - revisar',
        n=fresh_n.get(curso, 0),
    )
    current_order.append(curso)

# actualizar conteo de registros (n) para todos con datos frescos, aunque no haya cambiado la duracion
for curso in current_order:
    if curso in fresh_n:
        current[curso]['n'] = fresh_n[curso]

print('=== NUEVOS (%d) ===' % len(nuevos))
for c in nuevos:
    print(' +', c, '| duracion:', current[c]['dur'])
print()
print('=== RETIRADOS / ya no aparecen con "2026" (%d) ===' % len(retirados))
for c in retirados:
    print(' -', c)
print()
print('=== CAMBIOS DE DURACION NUMERICA (%d) ===' % len(cambios_num))
for c, old, new in cambios_num:
    print(f'  {c}: {old} -> {new} min')

with open('merge_report.json', 'w', encoding='utf-8') as f:
    json.dump(dict(nuevos=nuevos, retirados=retirados,
                    cambios=[[c, o, n] for c, o, n in cambios_num]),
              f, ensure_ascii=False, indent=1)

with open('merged_state.json', 'w', encoding='utf-8') as f:
    json.dump(dict(order=current_order, data=current), f, ensure_ascii=False, indent=1)
