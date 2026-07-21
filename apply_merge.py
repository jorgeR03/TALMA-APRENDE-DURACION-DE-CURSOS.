import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

FILE = 'Tiempo de Duracion de Cursos.xlsx'

with open('merged_state.json', encoding='utf-8') as f:
    merged = json.load(f)

rows = []
for curso in merged['order']:
    e = merged['data'][curso]
    rows.append(dict(curso=curso, inst=e['inst'] or '', dur=e['dur'], n=e['n'] or 0))

def tipo(dur):
    if isinstance(dur, (int, float)):
        return 'ok'
    if isinstance(dur, str):
        d = dur.strip().lower()
        if d == 'sin tiempo':
            return 'sin_tiempo'
        if d == 'sin examen':
            return 'sin_examen'
        if d.startswith('nuevo'):
            return 'nuevo'
    return 'otro'

for r in rows:
    r['tipo'] = tipo(r['dur'])

rows.sort(key=lambda r: r['curso'].upper())

n_total = len(rows)
n_ok = sum(1 for r in rows if r['tipo'] == 'ok')
n_sin_tiempo = sum(1 for r in rows if r['tipo'] == 'sin_tiempo')
n_sin_examen = sum(1 for r in rows if r['tipo'] == 'sin_examen')
n_nuevo = sum(1 for r in rows if r['tipo'] == 'nuevo')

buckets = [
    ('<= 15 min', 0, 15), ('16-30 min', 16, 30), ('31-60 min', 31, 60),
    ('61-120 min', 61, 120), ('121-240 min', 121, 240),
    ('241-480 min', 241, 480), ('> 480 min', 481, 10**9),
]
hist = [0] * len(buckets)
for r in rows:
    if r['tipo'] != 'ok':
        continue
    for i, (label, lo, hi) in enumerate(buckets):
        if lo <= r['dur'] <= hi:
            hist[i] += 1
            break

with open('duracion_dashboard_data.json', 'w', encoding='utf-8') as f:
    json.dump(dict(total=n_total, ok=n_ok, sin_tiempo=n_sin_tiempo, sin_examen=n_sin_examen, nuevo=n_nuevo,
                    buckets=[b[0] for b in buckets], hist=hist,
                    rows=[[r['curso'], r['inst'], (r['dur'] if r['tipo'] == 'ok' else None), r['tipo'], r['n']] for r in rows]),
              f, ensure_ascii=False, separators=(',', ':'))

# ---------------- Excel ----------------
NAVY = 'FF12233D'
NAVY_2 = 'FF1B3A63'
GRAY = 'FFF0F0EE'
GRAY_TXT = 'FF6B7280'
INFO_BG = 'FFE8F0FE'
INFO_TXT = 'FF2F6BE4'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Duracion Cursos'
ws.sheet_view.showGridLines = False

ws.merge_cells('A1:E2')
c = ws['A1']
c.value = '  TIEMPO DE DURACIÓN DE CURSOS — 2026'
c.font = Font(size=18, bold=True, color='FFFFFFFF')
c.alignment = Alignment(vertical='center')
for row_ in ws['A1:E2']:
    for cell in row_:
        cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 22

ws.merge_cells('A3:E3')
c = ws['A3']
c.value = (f'  Talma Servicios Aeroportuarios · {n_total} cursos con "2026" · {n_ok} con duración · '
           f'{n_sin_tiempo} sin tiempo · {n_sin_examen} sin examen · {n_nuevo} nuevos por revisar · '
           f'actualizado contra reporteglobal.xlsx del día')
c.font = Font(size=10, italic=True, color='FFFFFFFF')
c.alignment = Alignment(vertical='center')
for row_ in ws['A3:E3']:
    for cell in row_:
        cell.fill = PatternFill(start_color=NAVY_2, end_color=NAVY_2, fill_type='solid')
ws.row_dimensions[3].height = 18
ws.row_dimensions[4].height = 8

headers = ['CURSO', 'CLIENTE / INSTITUCIÓN', 'DURACIÓN (min)', 'DURACIÓN', 'REGISTROS EN 2026']
header_row = 5
for j, h in enumerate(headers):
    cell = ws.cell(row=header_row, column=1 + j, value=h)
    cell.font = Font(bold=True, color='FFFFFFFF', size=10)
    cell.fill = PatternFill(start_color=NAVY_2, end_color=NAVY_2, fill_type='solid')
    cell.alignment = Alignment(horizontal='center' if j > 0 else 'left', vertical='center')

thin = Side(style='thin', color='FFD8DCE3')
box = Border(left=thin, right=thin, top=thin, bottom=thin)

r = header_row + 1
for row_data in rows:
    minutos = row_data['dur'] if row_data['tipo'] == 'ok' else None
    if minutos is not None:
        hh, mm = divmod(int(minutos), 60)
        texto = f'{hh} h {mm} min' if hh and mm else (f'{hh} h' if hh else f'{mm} min')
    elif row_data['tipo'] == 'nuevo':
        texto = 'NUEVO - REVISAR'
    else:
        texto = row_data['dur']

    ws.cell(row=r, column=1, value=row_data['curso'])
    ws.cell(row=r, column=2, value=row_data['inst'])
    ws.cell(row=r, column=3, value=minutos if minutos is not None else texto)
    ws.cell(row=r, column=4, value=texto)
    ws.cell(row=r, column=5, value=row_data['n'])
    for col in range(1, 6):
        cell = ws.cell(row=r, column=col)
        cell.border = box
        if col >= 3:
            cell.alignment = Alignment(horizontal='center')
        if row_data['tipo'] in ('sin_tiempo', 'sin_examen'):
            cell.fill = PatternFill(start_color=GRAY, end_color=GRAY, fill_type='solid')
            if col in (3, 4):
                cell.font = Font(color=GRAY_TXT, italic=True)
        elif row_data['tipo'] == 'nuevo':
            cell.fill = PatternFill(start_color=INFO_BG, end_color=INFO_BG, fill_type='solid')
            if col in (3, 4):
                cell.font = Font(color=INFO_TXT, bold=True)
    r += 1
last_row = r - 1

ws.column_dimensions['A'].width = 62
ws.column_dimensions['B'].width = 34
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 16

tab = Table(displayName='DuracionCursos', ref=f'A{header_row}:E{last_row}')
tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
ws.add_table(tab)
ws.freeze_panes = f'A{header_row + 1}'

wb.save(FILE)
print('Excel actualizado:', FILE)
print('total:', n_total, '| con duracion:', n_ok, '| sin tiempo:', n_sin_tiempo,
      '| sin examen:', n_sin_examen, '| NUEVOS:', n_nuevo)
