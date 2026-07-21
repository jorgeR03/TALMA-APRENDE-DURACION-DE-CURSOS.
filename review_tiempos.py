import openpyxl, re
from collections import defaultdict

wb = openpyxl.load_workbook('Tiempo de Duracion de Cursos.xlsx', data_only=True)
ws = wb['Duracion Cursos']
rows = []
for r in range(6, ws.max_row + 1):
    curso = ws.cell(row=r, column=1).value
    if not curso:
        continue
    inst = ws.cell(row=r, column=2).value
    dur = ws.cell(row=r, column=3).value
    txt = ws.cell(row=r, column=4).value
    n = ws.cell(row=r, column=5).value
    rows.append(dict(curso=curso, inst=inst, dur=dur, txt=txt, n=n))

def is_num(v):
    return isinstance(v, (int, float))

# 1. Outliers: muy cortos o muy largos
print('=== Posibles outliers (muy cortos <10min o muy largos >480min) ===')
for r in rows:
    if is_num(r['dur']):
        if r['dur'] < 10 or r['dur'] > 480:
            print(f"{r['dur']:>5} min | {r['curso']} | n={r['n']}")

# 2. Familias: nombre base sin INICIAL/RECURRENTE/V2/_2026/BOG, comparar variantes
def base_name(c):
    c2 = c.upper()
    c2 = re.sub(r'\s*-?\s*(INICIAL|RECURRENTE)\s*', ' ', c2)
    c2 = re.sub(r'V\d+\b', ' ', c2)
    c2 = re.sub(r'_?\s*2026\s*BOG?', ' ', c2)
    c2 = re.sub(r'\s+', ' ', c2).strip()
    return c2

fam = defaultdict(list)
for r in rows:
    fam[base_name(r['curso'])].append(r)

print()
print('=== Familias con variantes de duracion inconsistentes (>2x entre variantes numericas) ===')
for name, items in fam.items():
    nums = [r['dur'] for r in items if is_num(r['dur'])]
    if len(nums) >= 2 and max(nums) > 0:
        ratio = max(nums) / max(min(nums), 1)
        if ratio >= 2.5:
            print(f"-- familia: {name}")
            for r in items:
                print(f"   {r['dur']} | {r['curso']}")

print()
print('=== Familias con mezcla de numerico y Sin tiempo/Sin examen (posible inconsistencia) ===')
for name, items in fam.items():
    types = set('num' if is_num(r['dur']) else r['dur'] for r in items)
    if len(items) >= 2 and len(types) > 1:
        print(f"-- familia: {name}")
        for r in items:
            print(f"   {r['dur']!r:>12} | {r['curso']}")
